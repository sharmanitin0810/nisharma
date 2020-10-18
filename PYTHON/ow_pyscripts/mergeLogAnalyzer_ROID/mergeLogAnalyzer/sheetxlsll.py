import time
import openpyxl
import xlrd
import csv
from openpyxl import load_workbook
import pandas as pd

def write_to_excel():
	print("Entering write to excel")
	wb = load_workbook(filename='UtBtHandoverSimulation.xlsx')
	ws = wb.worksheets[1]
	ws['B1'] = '10/08/2020  5:14:00 PM'
	#ws['B1'] = 'ONEWEB'
	time.sleep(4)
	wb.save(filename='UtBtHandoverSimulation.xlsx')
	time.sleep(4)
	df1 = pd.DataFrame(pd.read_excel(r"UtBtHandoverSimulation.xlsx")) 
	bool_series = pd.notnull(df1["Time"])  
	print(df1)
	print(bool_series)
	time.sleep(2)
	print("Going to Convert excel file to csv file")
	csv_from_excel()

def csv_from_excel():
	read_file = pd.read_excel (r'UtBtHandoverSimulation.xlsx',dtype=str, index_col=None)
	#read_file.replace('NaN','')
	read_file.to_csv (r'UtBtHandoverSimulation.csv', index = None, header=True , encoding = 'utf-8')
	#df = pd.DataFrame(pd.read_csv("UtBtHandoverSimulation.csv")) 
	
	#print(df['Time'])
	
if __name__ == "__main__":

	print("Entering main")
	write_to_excel()
